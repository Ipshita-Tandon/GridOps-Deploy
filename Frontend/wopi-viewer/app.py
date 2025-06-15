import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries, coordinate_from_string
from openpyxl.cell import Cell
from openpyxl.worksheet.views import SheetView, Selection, Pane
import sys
import os
import json

def highlight_cells(input_filepath, sheet_name, cell_addresses_json, merged_cells_data_json, charts_data_json, images_data_json, output_filepath):
    """
    Adds borders to multiple specified cells/ranges in an Excel file and saves a new copy.
    Determines merged cells, charts, and images based on the provided JSON data.
    Works with the specified sheet name (case-insensitive).
    Sets the active cell to the first highlighted cell to attempt focusing the view.
    """
    try:
        workbook = openpyxl.load_workbook(input_filepath)
        
        # Try to get the specified sheet (case-insensitive)
        sheet_name_lower = sheet_name.lower()
        actual_sheet_name = None
        for ws_name in workbook.sheetnames:
            if ws_name.lower() == sheet_name_lower:
                actual_sheet_name = ws_name
                break

        if actual_sheet_name is None:
            print(f"Error: Sheet '{sheet_name}' not found in workbook.", file=sys.stderr)
            print(f"Available sheets: {', '.join(workbook.sheetnames)}", file=sys.stderr)
            return 1

        worksheet = workbook[actual_sheet_name]
        print(f"Found sheet: {actual_sheet_name}", file=sys.stdout)

        # Make this sheet active in the workbook
        workbook.active = worksheet

        cell_addresses_to_process = json.loads(cell_addresses_json)
        merged_cells_data = json.loads(merged_cells_data_json)
        charts_data = json.loads(charts_data_json)
        images_data = json.loads(images_data_json)
        
        # Create border style with a nice blue color (RGB: 0, 120, 212)
        blue_side = Side(style='thick', color='0078D4')
        no_side = Side(style=None)

        processed_ranges_for_highlight = set()

        # Get all ranges from merged cells, charts, and images for the specified sheet
        all_special_ranges = []
        
        # Process merged cells (case-insensitive sheet name comparison)
        for sheet_info in merged_cells_data.get("merged_cells", []):
            if sheet_info.get("sheet_name", "").lower() == sheet_name_lower:
                for merged_cell in sheet_info.get("merged_cells", []):
                    all_special_ranges.append(merged_cell.get("range"))

        # Process charts - using the correct JSON structure (case-insensitive)
        for chart_info in charts_data.get("charts", []):
            if chart_info.get("sheet_name", "").lower() == sheet_name_lower:
                for chart in chart_info.get("charts_on_sheet", []):
                    # Handle both direct coords and series coords
                    if "coords" in chart:
                        all_special_ranges.append(chart["coords"])
                    # Also check series coords if they exist
                    for series in chart.get("series", []):
                        if "coords" in series:
                            all_special_ranges.append(series["coords"])

        # Process images - using the correct JSON structure (case-insensitive)
        for image_info in images_data.get("images", []):
            if image_info.get("sheet_name", "").lower() == sheet_name_lower:
                for image in image_info.get("images_on_sheet", []):
                    if "coords" in image:
                        all_special_ranges.append(image["coords"])

        if not all_special_ranges:
            print(f"Note: No special ranges found for sheet '{actual_sheet_name}' in provided JSON.", file=sys.stderr)

        first_highlighted_cell_coords = None # To store the first highlighted cell for setting active_cell

        for input_address in cell_addresses_to_process:
            input_address = input_address.strip().upper()

            if not input_address:
                continue

            if input_address in processed_ranges_for_highlight:
                continue

            cells_to_style_coords = []

            try:
                input_min_col, input_min_row, input_max_col, input_max_row = range_boundaries(input_address)
            except Exception as e:
                print(f"Warning: Invalid cell address or range '{input_address}' provided. Skipping: {e}", file=sys.stderr)
                continue

            # Check if this input address falls into any special range (merged cells, charts, or images)
            is_in_special_range = False
            for special_range in all_special_ranges:
                if not special_range:
                    continue

                try:
                    sr_min_col, sr_min_row, sr_max_col, sr_max_row = range_boundaries(special_range)
                except Exception as e:
                    print(f"Warning: Invalid special range '{special_range}' found in JSON. Skipping: {e}", file=sys.stderr)
                    continue

                if (input_min_row >= sr_min_row and input_max_row <= sr_max_row and
                    input_min_col >= sr_min_col and input_max_col <= sr_max_col):
                    cells_to_style_coords.append(range_boundaries(special_range))
                    processed_ranges_for_highlight.add(special_range)
                    is_in_special_range = True
                    print(f"Input '{input_address}' found within special range '{special_range}' in sheet '{actual_sheet_name}'. Highlighting entire range.", file=sys.stdout)
                    break

            if not is_in_special_range:
                cells_to_style_coords.append(range_boundaries(input_address))
                processed_ranges_for_highlight.add(input_address)
                print(f"Highlighting standalone cell/range: {input_address} in sheet '{actual_sheet_name}'", file=sys.stdout)

            # Apply borders to create a single box outline
            for min_col, min_row, max_col, max_row in cells_to_style_coords:
                # Store the first highlighted cell's coordinates
                if first_highlighted_cell_coords is None:
                    first_highlighted_cell_coords = (min_col, min_row)

                # For single cell, apply all borders
                if min_col == max_col and min_row == max_row:
                    cell = worksheet.cell(row=min_row, column=min_col)
                    cell.border = Border(
                        left=blue_side,
                        right=blue_side,
                        top=blue_side,
                        bottom=blue_side
                    )
                    continue

                # For ranges, apply borders to create a complete outline
                for row_idx in range(min_row, max_row + 1):
                    for col_idx in range(min_col, max_col + 1):
                        current_cell = worksheet.cell(row=row_idx, column=col_idx)
                        
                        # Get current cell's border
                        current_border = current_cell.border

                        # Determine which sides need borders
                        left = blue_side if col_idx == min_col else (current_border.left if current_border else no_side)
                        right = blue_side if col_idx == max_col else (current_border.right if current_border else no_side)
                        top = blue_side if row_idx == min_row else (current_border.top if current_border else no_side)
                        bottom = blue_side if row_idx == max_row else (current_border.bottom if current_border else no_side)

                        # Apply the complete border
                        current_cell.border = Border(
                            left=left,
                            right=right,
                            top=top,
                            bottom=bottom
                        )

        # Set up the sheet view with proper scroll position
        if first_highlighted_cell_coords:
            col, row = first_highlighted_cell_coords
            active_cell = f"{get_column_letter(col)}{row}"
            
            # Create a new sheet view if none exists
            if not worksheet.views:
                worksheet.views.append(SheetView())
            
            sheet_view = worksheet.views.sheetView[0]
            
            # Set zoom to 100%
            sheet_view.zoomScale = 100
            sheet_view.zoomScaleNormal = 100
            
            # Set the selection
            sheet_view.selection = [Selection(activeCell=active_cell, sqref=active_cell)]
            
            # Calculate scroll position (in rows and columns)
            # We want the highlighted cell to be visible near the top-left
            # but with some context around it
            scroll_row = max(1, row - 5)  # 5 rows above
            scroll_col = max(1, col - 2)  # 2 columns to the left
            
            # Set up the pane with split position to control scroll
            sheet_view.pane = Pane(
                xSplit=0,  # No horizontal split
                ySplit=0,  # No vertical split
                topLeftCell=f"{get_column_letter(scroll_col)}{scroll_row}",
                activePane="bottomRight",
                state="split"  # Use split instead of frozen
            )
            
            # Set the worksheet's scroll area
            worksheet.sheet_view.topLeftCell = f"{get_column_letter(scroll_col)}{scroll_row}"
            
            print(f"Setting active cell to: {active_cell} with scroll position at {worksheet.sheet_view.topLeftCell}", file=sys.stdout)
        else:
            print("No cells were highlighted, not setting active cell.", file=sys.stdout)

        # Save the modified workbook
        workbook.save(output_filepath)
        print(f"Successfully created bordered file: {output_filepath}")
        
        return 0, None # No image path in this version
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON provided: {e}", file=sys.stderr)
        return 1, None
    except Exception as e:
        print(f"Error processing Excel file: {e}", file=sys.stderr)
        return 1, None

if __name__ == "__main__":
    if len(sys.argv) != 8:
        print("Usage: python app.py <input_filepath> <sheet_name> <cell_addresses_json> <merged_cells_data_json> <charts_data_json> <images_data_json> <output_filepath>", file=sys.stderr)
        sys.exit(1)

    input_file = sys.argv[1]
    sheet_name = sys.argv[2]
    cell_addresses_json = sys.argv[3]
    merged_cells_data_json = sys.argv[4]
    charts_data_json = sys.argv[5]
    images_data_json = sys.argv[6]
    output_file = sys.argv[7]

    if not os.path.exists(input_file):
        print(f"Error: Input file not found at {input_file}", file=sys.stderr)
        sys.exit(1)

    return_code, _ = highlight_cells(input_file, sheet_name, cell_addresses_json, merged_cells_data_json, charts_data_json, images_data_json, output_file)
    
    sys.exit(return_code)
