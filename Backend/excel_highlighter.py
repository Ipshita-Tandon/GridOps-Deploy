"""
Excel Cell Highlighting Module

This module provides functionality to highlight specific cells/ranges in Excel files
with professional blue borders. Uses the exact implementation from server/app.py.
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries, coordinate_from_string
from openpyxl.cell import Cell
from openpyxl.worksheet.views import SheetView, Selection, Pane
import sys
import os
import json
import copy


def highlight_cells(workbook, sheet_name, cell_addresses_json, merged_cells_data_json, charts_data_json, images_data_json, output_filepath):
    """
    Adds borders to multiple specified cells/ranges in an Excel file and saves a new copy.
    Preserves all original formatting including background colors, charts, and images.
    Only modifies the border properties of specified cells.
    """
    try:
        # Try to get the specified sheet (case-insensitive)
        sheet_name_lower = sheet_name.lower()
        actual_sheet_name = None
        for ws_name in workbook.sheetnames:
            if ws_name.lower() == sheet_name_lower:
                actual_sheet_name = ws_name
                break

        if actual_sheet_name is None:
            print(f"Error: Sheet '{sheet_name}' not found in workbook.", file=sys.stderr)
            print(f"Available sheets: {', '.join(workbook.sheetnames)}", file=sys.stderr)
            return {'success': False, 'error': f"Sheet '{sheet_name}' not found in workbook."}

        worksheet = workbook[actual_sheet_name]
        print(f"Found sheet: {actual_sheet_name}", file=sys.stdout)

        # Make this sheet active in the workbook
        workbook.active = worksheet

        cell_addresses_to_process = json.loads(cell_addresses_json)
        merged_cells_data = json.loads(merged_cells_data_json)
        charts_data = json.loads(charts_data_json)
        images_data = json.loads(images_data_json)
        
        # Create border style with a nice blue color (RGB: 0, 120, 212)
        blue_side = Side(style='thick', color='0078D4')

        processed_ranges_for_highlight = set()

        # Get all ranges from merged cells, charts, and images for the specified sheet
        all_special_ranges = []

        # Process merged cells (case-insensitive sheet name comparison)
        for sheet_info in merged_cells_data.get("merged_cells", []):
            if sheet_info.get("sheet_name", "").lower() == sheet_name_lower:
                for merged_cell in sheet_info.get("merged_cells", []):
                    all_special_ranges.append(merged_cell.get("range"))

        # Process charts - using the correct JSON structure (case-insensitive)
        for sheet_info in charts_data.get("charts", []):
            if sheet_info.get("sheet_name", "").lower() == sheet_name_lower:
                for chart in sheet_info.get("charts_on_sheet", []):
                    if "coords" in chart:
                        all_special_ranges.append(chart["coords"])

        # Process images - using the correct JSON structure (case-insensitive)
        for sheet_info in images_data.get("images", []):
            if sheet_info.get("sheet_name", "").lower() == sheet_name_lower:
                for image in sheet_info.get("images_on_sheet", []):
                    if "coords" in image:
                        all_special_ranges.append(image["coords"])

        if not all_special_ranges:
            print(f"Note: No special ranges found for sheet '{actual_sheet_name}' in provided JSON.", file=sys.stderr)

        first_highlighted_cell_coords = None # To store the first highlighted cell for setting active_cell

        for input_address in cell_addresses_to_process:
            input_address = input_address.strip().upper()

            if not input_address:
                continue

            if input_address in processed_ranges_for_highlight:
                continue

            cells_to_style_coords = []

            try:
                input_min_col, input_min_row, input_max_col, input_max_row = range_boundaries(input_address)
            except Exception as e:
                print(f"Warning: Invalid cell address or range '{input_address}' provided. Skipping: {e}", file=sys.stderr)
                continue

            # Check if this input address falls into any special range (merged cells, charts, or images)
            is_in_special_range = False
            for special_range in all_special_ranges:
                if not special_range:
                    continue

                try:
                    sr_min_col, sr_min_row, sr_max_col, sr_max_row = range_boundaries(special_range)
                except Exception as e:
                    print(f"Warning: Invalid special range '{special_range}' found in JSON. Skipping: {e}", file=sys.stderr)
                    continue

                if (input_min_row >= sr_min_row and input_max_row <= sr_max_row and
                    input_min_col >= sr_min_col and input_max_col <= sr_max_col):
                    cells_to_style_coords.append(range_boundaries(special_range))
                    processed_ranges_for_highlight.add(special_range)
                    is_in_special_range = True
                    print(f"Input '{input_address}' found within special range '{special_range}' in sheet '{actual_sheet_name}'. Highlighting entire range.", file=sys.stdout)
                    break

            if not is_in_special_range:
                cells_to_style_coords.append(range_boundaries(input_address))
                processed_ranges_for_highlight.add(input_address)
                print(f"Highlighting standalone cell/range: {input_address} in sheet '{actual_sheet_name}'", file=sys.stdout)

            # Apply borders to create a single box outline while preserving all other formatting
            for min_col, min_row, max_col, max_row in cells_to_style_coords:
                # Store the first highlighted cell's coordinates
                if first_highlighted_cell_coords is None:
                    first_highlighted_cell_coords = (min_col, min_row)

                # For single cell, apply all borders while preserving existing formatting
                if min_col == max_col and min_row == max_row:
                    cell = worksheet.cell(row=min_row, column=min_col)
                    # Create new border while preserving existing cell properties
                    existing_border = cell.border
                    new_border = Border(
                        left=blue_side,
                        right=blue_side,
                        top=blue_side,
                        bottom=blue_side
                    )
                    cell.border = new_border
                    continue

                # For ranges, apply borders to create a complete outline while preserving formatting
                for row_idx in range(min_row, max_row + 1):
                    for col_idx in range(min_col, max_col + 1):
                        current_cell = worksheet.cell(row=row_idx, column=col_idx)
                        
                        # Create new border while preserving existing cell properties
                        new_border = Border(
                            left=blue_side,
                            right=blue_side,
                            top=blue_side,
                            bottom=blue_side
                        )
                        current_cell.border = new_border

        # Set up the sheet view with proper scroll position
        if first_highlighted_cell_coords:
            col, row = first_highlighted_cell_coords
            active_cell = f"{get_column_letter(col)}{row}"
            
            # Create a new sheet view if none exists
            if not worksheet.views:
                worksheet.views.append(SheetView())
            
            sheet_view = worksheet.views.sheetView[0]
            
            # Set zoom to 100%
            sheet_view.zoomScale = 100
            sheet_view.zoomScaleNormal = 100
            
            # Set the selection
            sheet_view.selection = [Selection(activeCell=active_cell, sqref=active_cell)]
            
            # Calculate scroll position (in rows and columns)
            scroll_row = max(1, row - 5)  # 5 rows above
            scroll_col = max(1, col - 2)  # 2 columns to the left
            
            # Set up the pane with split position to control scroll
            sheet_view.pane = Pane(
                xSplit=0,  # No horizontal split
                ySplit=0,  # No vertical split
                topLeftCell=f"{get_column_letter(scroll_col)}{scroll_row}",
                activePane="bottomRight",
                state="split"  # Use split instead of frozen
            )
            
            # Set the worksheet's scroll area
            worksheet.sheet_view.topLeftCell = f"{get_column_letter(scroll_col)}{scroll_row}"
            
            print(f"Setting active cell to: {active_cell} with scroll position at {worksheet.sheet_view.topLeftCell}", file=sys.stdout)
        else:
            print("No cells were highlighted, not setting active cell.", file=sys.stdout)

        # Save the modified workbook while preserving all formatting
        workbook.save(output_filepath)
        print(f"Successfully created bordered file: {output_filepath}")
        
        return {'success': True, 'error': None}
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON provided: {e}", file=sys.stderr)
        return {'success': False, 'error': f'Invalid JSON provided: {str(e)}'}
    except Exception as e:
        print(f"Error processing Excel file: {e}", file=sys.stderr)
        return {'success': False, 'error': f'Error processing Excel file: {str(e)}'}


def highlight_excel_cells(input_file_path, sheet_name, cell_ranges, output_file_path):
    """
    Wrapper function for the main highlighting functionality.
    Converts simple parameters to the format expected by highlight_cells function.
    Preserves all original Excel formatting.
    """
    try:
        # Load workbook with data_only=False to preserve all formatting
        workbook = openpyxl.load_workbook(input_file_path, data_only=False)
        
        # Prepare cell ranges as JSON array
        cell_ranges_array = [r.strip() for r in cell_ranges.split(',') if r.strip()]
        if not cell_ranges_array:
            return {'success': True, 'error': None}  # No cells to highlight
            
        cell_addresses_json = json.dumps(cell_ranges_array)
        
        # Get real data from the excel file using openpyxl
        worksheet = workbook[sheet_name]
        merged_cells_data = {
            "merged_cells": [
                {
                    "sheet_name": sheet_name,
                    "merged_cells": [
                        {
                            "range": merged_range.coord
                        } for merged_range in worksheet.merged_cells.ranges
                    ]
                }
            ]
        }
        
        # Get charts data
        charts_data = {
            "charts": [
                {
                    "sheet_name": sheet_name,
                    "charts_on_sheet": [
                        {
                            "coords": f"{get_column_letter(chart.anchor._from.col)}{chart.anchor._from.row}"
                        } for chart in worksheet._charts
                    ]
                }
            ]
        }
        
        # Get images data
        images_data = {
            "images": [
                {
                    "sheet_name": sheet_name,
                    "images_on_sheet": [
                        {
                            "coords": f"{get_column_letter(img.anchor._from.col)}{img.anchor._from.row}"
                        } for img in worksheet._images
                    ]
                }
            ]
        }
        
        # Convert to JSON strings
        merged_cells_json = json.dumps(merged_cells_data)
        charts_json = json.dumps(charts_data)
        images_json = json.dumps(images_data)
        
        # Call the main highlighting function
        return highlight_cells(
            workbook,
            sheet_name,
            cell_addresses_json,
            merged_cells_json,
            charts_json,
            images_json,
            output_file_path
        )
    except Exception as e:
        print(f"Error in highlight_excel_cells: {e}", file=sys.stderr)
        return {'success': False, 'error': f'Error in highlight_excel_cells: {str(e)}'} 