from openpyxl import load_workbook
import os
import json
import zipfile

# Basic cell content extraction text, formula, result, hyperlinks, coordinates
def extract_cell_content(file_path, output_path):
    pass
# Style info extraction font, color, fill
def extract_style_info():
    pass

# Image extraction
def extract_images():

    pass

# Chart extraction
def extract_charts():
    pass

########################################################
# FUNCTION DEFINITIONS
########################################################
def extract_cell_content(file_path, output_path):
    # Load workbooks - one for formulas and one for values
    wb_vals = load_workbook(file_path, data_only=True)  # For cell values
    wb_formulas = load_workbook(file_path, data_only=False)  # For formulas
    
    # Dictionary to store all sheet data with schema
    all_sheets_data = {
        "schema": ["value", "formula", "hyperlink"]
    }
    
    # Process each sheet
    for sheet_name in wb_vals.sheetnames:
        ws_vals = wb_vals[sheet_name]
        ws_formulas = wb_formulas[sheet_name]
        
        # Use dictionary instead of nested arrays for direct cell access
        sheet_data = {}
        
        # First process merged cells
        for merged_range in ws_vals.merged_cells.ranges:
            # Get the value from the top-left cell of the merged range
            top_left_cell = ws_vals[merged_range.start_cell.coordinate]
            if top_left_cell.value is not None:
                value = str(top_left_cell.value)
                formula = str(ws_formulas[merged_range.start_cell.coordinate].value) if (ws_formulas[merged_range.start_cell.coordinate].value is not None and ws_formulas[merged_range.start_cell.coordinate].value != top_left_cell.value) else None
                sheet_data[merged_range.coord] = [value, formula]
        
        # Then process remaining non-merged cells
        for row in ws_vals.iter_rows():
            for cell in row:
                # Skip if cell is part of a merged range
                if not any(cell.coordinate in merged_range for merged_range in ws_vals.merged_cells.ranges):
                    if cell.value is not None:  # Only process non-empty cells
                        value = str(cell.value)
                        formula = str(ws_formulas[cell.coordinate].value) if (ws_formulas[cell.coordinate].value is not None and ws_formulas[cell.coordinate].value != cell.value) else None
                        sheet_data[cell.coordinate] = [value, formula]
        
        all_sheets_data[sheet_name] = sheet_data
    
    # Write to JSON file
    file_name = os.path.basename(file_path).split(".")[0]
    output_file = os.path.join(output_path, f"{file_name}_cell_content.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_sheets_data, f, indent=2, ensure_ascii=False)
    
    print(f"Cell content extracted and saved to {output_file}")  

def main():
    # input_path = input("Enter the path to the Excel file: ")
    # output_path = input("Enter the path to the output directory: ")
    input_path = "/Users/yashwanthsaip/Desktop/Excel-Project/Backend/Openpyxl/Excel_files/Annual Financial Report1.xlsx"
    output_path = "/Users/yashwanthsaip/Desktop/Excel-Project/Backend/Openpyxl/extract-output"
    if not os.path.exists(input_path):
        print(f"Error: Input file does not exist: {input_path}")
        return
    
    if not os.path.exists(output_path):
        os.makedirs(output_path)

    extract_cell_content(input_path, output_path)

if __name__ == "__main__":
    main()
        
    
        